import os
from openpyxl import load_workbook
import grig_module
import ml

def count_non_empty_columns(xlsx_file):
    # Загружаем Excel файл
    workbook = load_workbook(xlsx_file)
    non_empty_column_count = 0

    # Проходим по всем листам в книге
    for sheet in workbook.worksheets:
        # Получаем максимальное количество колонок в листе
        max_column = sheet.max_column

        # Проверяем каждую колонку
        for col in range(1, max_column + 1):
            # Проверяем, есть ли непустые ячейки в колонке
            # Используем генератор для проверки наличия непустых ячеек
            column_values = sheet.iter_cols(min_col=col, max_col=col, values_only=True)
            if any(cell is not None for cell in next(column_values)):
                non_empty_column_count += 1

    return non_empty_column_count

def ml_processing(xlsx_file):
    folder_name = "static"
    questions, clusters = ml.process(xlsx_file)
    tags = grig_module.get_tags(clusters)
    texts = grig_module.get_texts(clusters)
    for i, text in enumerate(texts):
        with open(os.path.join(folder_name, f"text{i + 1}.txt"), 'w') as f:
            f.write(text)
    picture_names = grig_module.get_pictures(tags, folder_name)

def process(xlsx_file):
    n = count_non_empty_columns(xlsx_file)
    # Задаем папку для сохранения .txt файлов
    output_folder = 'static'
    os.makedirs(output_folder, exist_ok=True)

    ml_processing(xlsx_file)
